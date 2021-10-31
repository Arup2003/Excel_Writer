import openpyxl
import time
from tkinter import *
from tkinter import filedialog, ttk
from tkinter import messagebox
import threading

choose = "dark"

window = Tk()
window.title("Arup's Excel Writer")
window.configure(background = "#01723A")
window.geometry('1000x700')
window.resizable(width = False, height = False)

def mode_switch():
	global choose
	if choose == "dark":
		background.config(image = light)
		choose = "light"
		DarkMode.config(image = button_light)
		Read_File_name_entry.config(bg = "#EBF5F5", fg = "brown")
		Write_File_name_entry.config(bg = "#EBF5F5", fg = "brown")
		Read_File_ROW_from_entry.config(bg = "#EBF5F5", fg = "brown")
		Read_File_ROW_to_entry.config(bg = "#EBF5F5", fg = "brown")
		Read_File_COLUMN_from_entry.config(bg = "#EBF5F5", fg = "brown")
		Read_File_COLUMN_to_entry.config(bg = "#EBF5F5", fg = "brown")
		Column_Fixed.config(bg = "#EBF5F5", fg = "white")
		Row_Fixed.config(bg = "#EBF5F5", fg = "white")
		None_Type_entry.config(bg = "#EBF5F5", fg = "brown")
		Write_File_fixed_entry.config(bg = "#EBF5F5", fg = "brown")
		Write_File_write_entry.config(bg = "#EBF5F5", fg = "brown")
	else: 
		background.config(image = dark)
		choose = "dark"
		DarkMode.config(image = button_dark)
		Read_File_name_entry.config(bg = "#242424", fg = "#D4CFCF")
		Write_File_name_entry.config(bg = "#242424", fg = "#D4CFCF")
		Read_File_ROW_from_entry.config(bg = "#242424", fg = "#D4CFCF")
		Read_File_ROW_to_entry.config(bg = "#242424", fg = "#D4CFCF")
		Read_File_COLUMN_from_entry.config(bg = "#242424", fg = "#D4CFCF")
		Read_File_COLUMN_to_entry.config(bg = "#242424", fg = "#D4CFCF")
		Column_Fixed.config(bg = "#242424", fg = "#D4CFCF")
		Row_Fixed.config(bg = "#242424", fg = "#D4CFCF")
		None_Type_entry.config(bg = "#242424", fg = "#D4CFCF")
		Write_File_fixed_entry.config(bg = "#242424", fg = "#D4CFCF")
		Write_File_write_entry.config(bg = "#242424", fg = "#D4CFCF")




def prog():
	if int(R_C_Loop.get()) == 0:
		messagebox.showinfo("Incomplete", "Please select if row or column is fixed")
	elif Read_File_name_entry.get() == "" or Write_File_name_entry == "":
		messagebox.showinfo("Incomplete", "Please specify File name")
	elif Read_File_ROW_from_entry.get() == "" or Write_File_fixed_entry.get()== "" or Write_File_write_entry.get() == ""or Read_File_ROW_to_entry.get() == ""or Read_File_COLUMN_from_entry.get() == ""or Read_File_COLUMN_to_entry.get()== "":
		messagebox.showinfo("Incomplete", "Please fill every specifications")
	
	else:
		global pb
		pb = ttk.Progressbar(orient = "horizontal", length = 425, mode = "determinate")
		pb.place(x=300,y= 390)
		if choose == "dark":
			background.config(image = dark_blur)
		else:
			background.config(image = light_blur)
		Read_File_name_entry.place(x = 1000, y = 1000)
		Write_File_name_entry.place(x =1000,y=1999)
		Read_File_ROW_from_entry.place(x=1000,y=1000)
		Start.place(x=1200,y=4924)
		Write_File_fixed_entry.place(x = 5519, y = 6130)

		Write_File_write_entry.place(x = 5160, y = 5110)
		Row_Fixed.place(x = 8215, y = 3310)
		Column_Fixed.place(x = 8251, y = 3179)
		None_Type_entry.place(x = 1105, y = 6135)
		Read_File_ROW_to_entry.place(x=10100,y=4101)
		Read_File_COLUMN_from_entry.place(x=249492,y=4929)
		Read_File_COLUMN_to_entry.place(x=4295,y=2425)
		Start1.place(x = 1000, y = 1000)
		DarkMode.place(x = 1000, y = 2000)
		window.update()
		program()


def browse():
	window.filename = filedialog.askopenfilename(title = "Select a file")
	Read_File_name_entry.insert(0, window.filename)

alphabets = ["a","b","c","d","e","f","g","h","i","j","k",'l','m','n','o','p','q','r','s','t','u','v','w','x','y','z','aa','ab','ac','ad','ae','af','ag','ah','ai','aj','ak','al','am','an','ao','ap','aq','ar', "as", "at","au", "av", "aw", "ax", "ay", "az"]

def program():
	file_num = 1
	write = int(Write_File_write_entry.get())
	fixed = int(Write_File_fixed_entry.get())
	Read_File = Read_File_name_entry.get()
	Write_File = Write_File_name_entry.get()
	read_row_from = Read_File_ROW_from_entry
	read_row_to = Read_File_ROW_to_entry
	roll_nos_from = int(Read_File_ROW_from_entry.get())
	roll_nos_to = int(Read_File_ROW_to_entry.get()) + 1
	read_column_from = (Read_File_COLUMN_from_entry.get())
	read_column_to = (Read_File_COLUMN_to_entry.get())
	No_Entry = None_Type_entry.get()

	if read_column_from.lower() in alphabets:
		read_column_from = int(alphabets.index(read_column_from.lower()) + 1)
	if read_column_to.lower() in alphabets:
		read_column_to = int(alphabets.index(read_column_to.lower()) + 2)


	
	book = openpyxl.load_workbook(Read_File, data_only =True)

	sheet = book.active
	

	#-------------------------------ROW RANGE IN READ FILE------------------------------------#
	for row in range(roll_nos_from, roll_nos_to):
	#-----------------------------------------------------------------------------------------#

	#-----------------------------COLUMN RANGE IN READ FILE-----------------------------------#
		for column in range(read_column_from, read_column_to):
	#-----------------------------------------------------------------------------------------#

			a = sheet.cell(row = row, column = column)
			grade = a.value

	#==============================  WHAT TO DO WHEN THE CELL IS BLANK =========================================#
			if grade == None:
				grade = No_Entry
			print(grade)
	#===========================================================================================================#
			
			#This file will be opened to write data in (You can also make a new file)
			file = openpyxl.load_workbook("Files\\"+Write_File+str(file_num)+'.xlsx')
			indi_doc = file.active

	#===========================ROW AND COLUMN OF WRITE FILE (write variable is the starting cell)============================#
			if int(R_C_Loop.get()) == 2:
				indi_doc.cell(row = write , column = fixed).value = grade

			elif int(R_C_Loop.get()) == 1:
				indi_doc.cell(row = fixed , column = write).value = grade
			
			else:
				print("OH NO!")
				print(int(R_C_Loop.get()))

	#==========================================================================================================================#
			#This will be the name of output file
			file.save("Files\\"+Write_File+str(file_num)+'.xlsx')		
			write +=1
		pb['value'] = (file_num/(roll_nos_to - roll_nos_from))*100
		window.update()
		print(Write_File+str(file_num)+'.xlsx'+" Completed...")  
		
		write = int(Write_File_write_entry.get())
		file_num +=1
	
	time.sleep(3)
	if choose == "dark":
		background.config(image = dark)
	else: 
		background.config(image = light)

	DarkMode.place(x= 0, y = 0)
	Start1.place(x = 830, y= 13)	
	Read_File_name_entry.place(x = 110, y = 250)
	Read_File_name_entry.delete(0,'end')
	Write_File_name_entry.place(x =565 ,y=250)
	Write_File_name_entry.delete(0,'end')
	Read_File_ROW_from_entry.place(x=115,y=380)
	Read_File_ROW_from_entry.delete(0,'end')
	Start.place(x=420,y=250)
	Write_File_fixed_entry.place(x = 559, y = 630)
	Write_File_fixed_entry.delete(0,'end')
	Write_File_write_entry.place(x = 560, y = 510)
	Write_File_write_entry.delete(0,'end')
	Row_Fixed.place(x = 825, y = 330)
	Column_Fixed.place(x = 825, y = 379)
	None_Type_entry.place(x = 105, y = 635)
	None_Type_entry.delete(0,'end')
	Read_File_ROW_to_entry.place(x=320,y=380)
	Read_File_ROW_to_entry.delete(0,'end')
	Read_File_COLUMN_from_entry.place(x=115,y=515)
	Read_File_COLUMN_from_entry.delete(0,'end')
	Read_File_COLUMN_to_entry.place(x=325,y=515)
	Read_File_COLUMN_to_entry.delete(0,'end')
	pb.destroy()

	print("Every file written successfully")
	window.mainloop()

#_________________________________________________________________________________________________________________________________
img = PhotoImage(file = "images/yo.png")
dark = PhotoImage(file = "images/dark.png")
light = PhotoImage(file = "images/light.png")
button_dark = PhotoImage(file = "images/dark_mode.png")
button_light = PhotoImage(file = "images/light_mode.png")
dark_blur = PhotoImage(file = "images/dark_blur.png")
light_blur = PhotoImage(file = "images/light_blur.png")

R_C_Loop = IntVar()

background = Label(image = dark)
background.place(x=0,y=0)

Read_File_name_entry = Entry(window, font ='Calibri, 20', state = 'normal', borderwidth = 0, bg = "#242424",fg = "#D4CFCF")
Read_File_name_entry.place(x = 110, y = 255)

Read_File_ROW_from_entry = Entry(window, width = 10, font ='Calibri, 20', state = 'normal', borderwidth = 0,bg = "#242424",fg = "#D4CFCF")
Read_File_ROW_from_entry.place(x = 105, y = 390)
Read_File_ROW_to_entry = Entry(window, width = 10, font ='Calibri, 20', state = 'normal',borderwidth = 0,bg = "#242424",fg = "#D4CFCF")
Read_File_ROW_to_entry.place(x = 305, y = 390)

Read_File_COLUMN_from_entry = Entry(window, width = 10, font ='Calibri, 20', state = 'normal', borderwidth = 0,bg = "#242424", fg = "#D4CFCF")
Read_File_COLUMN_from_entry.place(x = 110, y = 520)
Read_File_COLUMN_to_entry = Entry(window, width = 10, font ='Calibri, 20', state = 'normal', borderwidth = 0,bg = "#242424",fg = "#D4CFCF")
Read_File_COLUMN_to_entry.place(x=305, y = 520)

None_Type_entry = Entry(window, width = 10, font ='Calibri, 20', state = 'normal', borderwidth = 0,bg = "#242424",fg = "#D4CFCF")
None_Type_entry.place(x = 105, y = 635)

Write_File_name_entry = Entry(window, font ='Calibri, 20', state = 'normal',borderwidth = 0,bg = "#242424",fg = "#D4CFCF")
Write_File_name_entry.place(x= 565, y = 255)

Row_Fixed = Radiobutton(window, variable = R_C_Loop, bg = "#242424", value = 1, font = ("Bahnschrift, 15"), selectcolor=  "black", foreground = "#D4CFCF")
Column_Fixed = Radiobutton(window, variable = R_C_Loop, bg = "#242424", value = 2, font = ("Bahnschrift, 15"), foreground = "#D4CFCF", selectcolor = "black")
Row_Fixed.place(x = 825, y = 330)
Column_Fixed.place(x = 825, y = 379)

Write_File_write_entry = Entry(window, width = 10, font ='Calibri, 20', state = 'normal', borderwidth = 0,bg = "#242424",fg = "#D4CFCF")
Write_File_write_entry.place(x = 560, y = 520)

Write_File_fixed_entry = Entry(window, width = 10, font ='Calibri, 20', state = 'normal', borderwidth = 0,bg = "#242424",fg = "#D4CFCF")
Write_File_fixed_entry.place(x = 559, y = 630)

Start = Button(command = lambda:browse(), borderwidth = 0.6, text = "Browse",font =("Product Sans", "15"), foreground = "#7DBD4C", background = "#101010")
Start1 = Button(image = img,command = lambda:prog(), background = "#2EA04B", borderwidth = 1)
DarkMode = Button(image = button_dark, command = lambda:mode_switch(), background = "#EBF5F5", height = 83, width = 79, bd = 0)
Start.place(x = 420, y = 250)
Start1.place(x = 830, y = 13)
DarkMode.place(x=0, y = 0)

window.mainloop()
